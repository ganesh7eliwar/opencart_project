import openpyxl
from openpyxl.styles import PatternFill


def total_rows(file_name, sheet_name):
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook[sheet_name]

    return sheet.max_row


def total_columns(file_name, sheet_name):
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook[sheet_name]

    return sheet.max_column


def read_data(file_name, sheet_name, row_number, column_number):
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook[sheet_name]

    return sheet.cell(row_number, column_number).value


def write_data(file_name, sheet_name, row_number, column_number, data):
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook[sheet_name]
    sheet.cell(row_number, column_number).value = data
    workbook.save(file_name)


def green_color(file_name, sheet_name, row_number, column_number):
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook[sheet_name]
    fill_green = PatternFill(start_color='79DB77', end_color='79DB77', fill_type='solid')
    sheet.cell(row_number, column_number).fill = fill_green
    workbook.save(file_name)


def red_color(file_name, sheet_name, row_number, column_number):
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook[sheet_name]
    fill_red = PatternFill(start_color='F76143', end_color='F76143', fill_type='solid')
    sheet.cell(row_number, column_number).fill = fill_red
    workbook.save(file_name)
